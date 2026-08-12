"""
Planilha de upgrade das marcas de plano antigo ("Mensalidade ate 400 .xlsx").

Dois passos:

  py upgrade_planilha.py preparar
      Le as 3 abas do xlsx original, consolida numa unica aba e JA PRE-PREENCHE
      o que da' pra tirar da base da Vesti (CNPJ, razao social, nº de filiais,
      GMV do trimestre, mensalidade, CS, canal) cruzando pelo nome da marca.
      Gera: upgrade_marcas.xlsx  -> voce completa as colunas CNPJ e Instagram
      onde ficaram vazias e roda o passo 2.

  py upgrade_planilha.py enriquecer
      Le upgrade_marcas.xlsx, consulta cada CNPJ na BrasilAPI (Receita Federal,
      de graca, sem chave) e preenche CNAE, porte, matriz/filial, situacao e
      endereco. Com isso calcula "provavel loja fisica" + o motivo.
      Seguidores do Instagram: so' se INSTAGRAM_API_KEY estiver setada (RapidAPI);
      senao a coluna fica vazia pra preencher a mao. Ver README_UPGRADE.md.

As respostas da Receita ficam em cache (.cache_cnpj.json) -- rodar de novo nao
refaz as consultas ja' feitas.
"""
from __future__ import annotations

import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
XLSX_ORIGEM = ROOT / "Mensalidade até 400 .xlsx"
XLSX_SAIDA = ROOT / "upgrade_marcas.xlsx"
CACHE_CNPJ = ROOT / ".cache_cnpj.json"
DASHBOARD = ROOT / "dashboard_data.js"

# CNAEs que indicam venda em LOJA FISICA (comercio varejista em estabelecimento)
CNAE_LOJA = {"4781", "4782", "4783", "4789", "4755", "4713", "4712"}
# CNAEs de venda NAO presencial (internet / catalogo / porta a porta)
CNAE_ONLINE = {"4791", "4792", "4793"}
# Industria/confeccao: nao diz nada sobre loja, mas indica fabricante
CNAE_INDUSTRIA = {"1411", "1412", "1413", "1421", "1422", "1340", "1330"}


# ---------------------------------------------------------------- utilidades
def _norm(s) -> str:
    """Normaliza nome de marca pra casar entre planilha e base (sem acento/pontuacao)."""
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(ltda|me|epp|eireli|sa|s/a|comercio|confeccoes|moda[s]?)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _so_digitos(s) -> str:
    return re.sub(r"\D", "", str(s or ""))


def _carrega_base_vesti() -> dict:
    """Indexa as marcas do painel por nome normalizado."""
    if not DASHBOARD.exists():
        print(f"[aviso] {DASHBOARD.name} nao encontrado -- sem pre-preenchimento da base.")
        return {}
    txt = DASHBOARD.read_text(encoding="utf-8")
    dados = json.loads(txt[txt.index("=") + 1:].rstrip().rstrip(";"))
    idx = {}
    for e in dados.get("empresas", []):
        for chave in {_norm(e.get("name")), _norm(e.get("nome_fantasia"))}:
            if chave and chave not in idx:
                idx[chave] = e
    return idx


def _gmv_trimestre(e: dict, meses: list[str]) -> float:
    return sum((e.get("mensal") or {}).get(m, {}).get("valTotal", 0) or 0 for m in meses)


def _ultimos3_fechados(dados_meses: list[str]) -> list[str]:
    from datetime import date
    hoje = date.today()
    atual = f"{hoje.year:04d}-{hoje.month:02d}"
    return sorted([m for m in dados_meses if m < atual])[-3:]


# ---------------------------------------------------------------- passo 1
COLUNAS = [
    "Aba origem", "Marca", "Plano", "Mensalidade planilha", "CS planilha",
    "CNPJ", "Instagram (@)",                       # <- Laura completa o que faltar
    "CNPJ origem", "Razao social", "Marca na base?", "CS base", "Canal",
    "Mensalidade base", "Filiais na Vesti", "GMV 3 meses",
    "Seguidores Instagram",                        # <- passo 2 (se tiver chave)
    "Provavel loja fisica", "Motivo loja fisica",  # <- passo 2
    "CNAE principal", "CNAE descricao", "Porte", "Matriz/Filial",
    "Situacao cadastral", "Municipio", "UF",
]


def _linhas_do_xlsx_origem() -> list[dict]:
    """Le as 3 abas. Os cabecalhos sao inconsistentes (a aba Light/Starter nem tem),
    entao vamos por posicao de coluna, que e' estavel nas tres."""
    wb = openpyxl.load_workbook(XLSX_ORIGEM, data_only=True)
    out = []
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or not any(v not in (None, "") for v in row):
                continue
            marca = str(row[0] or "").strip()
            if not marca:
                continue
            # pula linha de cabecalho (so nas abas que tem)
            if _norm(marca) == "marca":
                continue
            plano = str(row[2] or "").strip() if len(row) > 2 else ""
            mensalidade = row[6] if len(row) > 6 else None
            cs = str(row[7] or "").strip() if len(row) > 7 else ""
            out.append({"aba": ws.title.strip(), "marca": marca, "plano": plano,
                        "mensalidade": mensalidade, "cs": cs})
    return out


def preparar() -> None:
    linhas = _linhas_do_xlsx_origem()
    base = _carrega_base_vesti()
    meses = []
    if DASHBOARD.exists():
        txt = DASHBOARD.read_text(encoding="utf-8")
        dados = json.loads(txt[txt.index("=") + 1:].rstrip().rstrip(";"))
        meses = _ultimos3_fechados(dados.get("mesesList", []))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marcas upgrade"
    ws.append(COLUNAS)

    # a aba "Vesti" e' superset da aba "CS ": deduplica por marca e guarda em
    # quais abas ela apareceu, senao a mesma marca vira 2 linhas pra ligar.
    por_marca: dict[str, dict] = {}
    for ln in linhas:
        k = _norm(ln["marca"])
        if not k:
            continue
        if k in por_marca:
            abas = por_marca[k]["abas"]
            if ln["aba"] not in abas:
                abas.append(ln["aba"])
            # completa campo faltante com o que a outra aba tiver
            for campo in ("plano", "mensalidade", "cs"):
                if not por_marca[k][campo] and ln[campo]:
                    por_marca[k][campo] = ln[campo]
            continue
        por_marca[k] = {**ln, "abas": [ln["aba"]]}

    achadas = 0
    vistas = por_marca
    for ln in por_marca.values():
        e = base.get(_norm(ln["marca"]))
        if e:
            achadas += 1
        ws.append([
            " + ".join(ln["abas"]), ln["marca"], ln["plano"], ln["mensalidade"], ln["cs"],
            _so_digitos(e.get("cnpj")) if e else "",          # CNPJ (pre-preenchido)
            "",                                                # Instagram (@) -> Laura
            "base Vesti" if (e and e.get("cnpj")) else "",
            (e or {}).get("razao_social", ""),
            "sim" if e else "NAO ENCONTRADA",
            (e or {}).get("cs", ""),
            (e or {}).get("canal", ""),
            (e or {}).get("valor_plano") or (e or {}).get("valor_mensal") or "",
            (e or {}).get("qtdFiliais", "") if e else "",
            round(_gmv_trimestre(e, meses), 2) if e else "",
        ] + [""] * (len(COLUNAS) - 15))

    # formatacao
    fill = PatternFill("solid", fgColor="6C5CE7")
    fill_edit = PatternFill("solid", fgColor="F39C12")
    for c in range(1, len(COLUNAS) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color="FFFFFF", size=9)
        cel.fill = fill_edit if COLUNAS[c - 1] in ("CNPJ", "Instagram (@)") else fill
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(30, len(COLUNAS[c - 1]) + 6))
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_SAIDA)

    total = len(vistas)
    sem_cnpj = total - sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[5])
    print(f"[ok] {XLSX_SAIDA.name}: {total} marcas ({achadas} casadas com a base da Vesti).")
    print(f"     Colunas em LARANJA sao as suas: CNPJ ({sem_cnpj} ainda vazias) e Instagram (@).")
    print(f"     Depois de preencher, rode:  py {Path(__file__).name} enriquecer")


# ---------------------------------------------------------------- passo 2
def _http_json(url: str, headers: dict | None = None, timeout: int = 25):
    req = urllib.request.Request(url, headers=headers or {"User-Agent": "painel-vesti/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8"))


def _consulta_cnpj(cnpj: str, cache: dict) -> dict | None:
    """BrasilAPI (Receita Federal). Gratis e sem chave, mas com rate limit."""
    if cnpj in cache:
        return cache[cnpj]
    url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
    for tentativa in range(4):
        try:
            data = _http_json(url)
            cache[cnpj] = data
            return data
        except urllib.error.HTTPError as ex:
            if ex.code == 429:                      # rate limit -> espera e tenta de novo
                time.sleep(8 * (tentativa + 1))
                continue
            cache[cnpj] = {"_erro": f"HTTP {ex.code}"}
            return cache[cnpj]
        except Exception as ex:                     # rede instavel
            if tentativa == 3:
                return {"_erro": str(ex)[:80]}
            time.sleep(3)
    return {"_erro": "429 apos varias tentativas"}


def _avalia_loja_fisica(rec: dict | None, filiais_vesti) -> tuple[str, str]:
    """Combina os sinais disponiveis. E' indicacao, nao certeza -- por isso o motivo."""
    motivos, pontos = [], 0
    try:
        n_fil = int(filiais_vesti or 0)
    except (TypeError, ValueError):
        n_fil = 0
    if n_fil >= 1:
        pontos += 2
        motivos.append(f"{n_fil} filial(is) cadastradas na Vesti")

    if rec and not rec.get("_erro"):
        principal = str(rec.get("cnae_fiscal") or "")[:4]
        secundarios = {str(c.get("codigo") or "")[:4] for c in (rec.get("cnaes_secundarios") or [])}
        todos = {principal} | secundarios
        if principal in CNAE_LOJA:
            pontos += 2
            motivos.append(f"CNAE principal {principal} = varejo em loja")
        elif todos & CNAE_LOJA:
            pontos += 1
            motivos.append(f"CNAE secundario de varejo em loja ({sorted(todos & CNAE_LOJA)[0]})")
        if principal in CNAE_ONLINE:
            pontos -= 2
            motivos.append(f"CNAE principal {principal} = venda pela internet/catalogo")
        if principal in CNAE_INDUSTRIA:
            motivos.append(f"CNAE principal {principal} = confeccao/industria")
        if str(rec.get("descricao_identificador_matriz_filial") or "").upper().startswith("FILIAL"):
            pontos += 1
            motivos.append("CNPJ e' FILIAL (grupo com mais de um estabelecimento)")
    elif not motivos:
        return ("sem dados", "sem CNPJ valido consultado")

    if pontos >= 3:
        veredito = "provavel SIM"
    elif pontos >= 1:
        veredito = "talvez"
    elif pontos <= -1:
        veredito = "provavel NAO"
    else:
        veredito = "indefinido"
    return (veredito, "; ".join(motivos) or "sem sinal claro")


def _seguidores_instagram(user: str) -> str:
    """So funciona com uma chave de API (RapidAPI). Sem chave, devolve vazio --
    o Instagram bloqueia requisicao anonima, entao nao ha caminho gratuito confiavel."""
    chave = os.environ.get("INSTAGRAM_API_KEY")
    host = os.environ.get("INSTAGRAM_API_HOST", "instagram-scraper-api2.p.rapidapi.com")
    if not chave or not user:
        return ""
    user = user.lstrip("@").strip()
    try:
        data = _http_json(
            f"https://{host}/v1/info?username_or_id_or_url={user}",
            headers={"x-rapidapi-key": chave, "x-rapidapi-host": host,
                     "User-Agent": "painel-vesti/1.0"})
        d = data.get("data") or data
        for k in ("follower_count", "followers", "edge_followed_by"):
            v = d.get(k)
            if isinstance(v, dict):
                v = v.get("count")
            if isinstance(v, (int, float)):
                return str(int(v))
    except Exception as ex:
        return f"erro: {str(ex)[:40]}"
    return ""


def enriquecer() -> None:
    if not XLSX_SAIDA.exists():
        sys.exit(f"[erro] {XLSX_SAIDA.name} nao existe. Rode primeiro: py {Path(__file__).name} preparar")
    cache = json.loads(CACHE_CNPJ.read_text(encoding="utf-8")) if CACHE_CNPJ.exists() else {}
    wb = openpyxl.load_workbook(XLSX_SAIDA)
    ws = wb.active
    cab = [c.value for c in ws[1]]
    col = {nome: i + 1 for i, nome in enumerate(cab)}

    tem_chave = bool(os.environ.get("INSTAGRAM_API_KEY"))
    if not tem_chave:
        print("[aviso] INSTAGRAM_API_KEY nao setada -> coluna de seguidores fica vazia.")

    consultados = com_loja = 0
    for r in range(2, ws.max_row + 1):
        cnpj = _so_digitos(ws.cell(row=r, column=col["CNPJ"]).value)
        filiais = ws.cell(row=r, column=col["Filiais na Vesti"]).value
        rec = None
        if len(cnpj) == 14:
            rec = _consulta_cnpj(cnpj, cache)
            consultados += 1
            if consultados % 10 == 0:
                CACHE_CNPJ.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
                print(f"  ... {consultados} CNPJs consultados")
            time.sleep(1.2)                     # respeita o rate limit da BrasilAPI

        if rec and not rec.get("_erro"):
            ws.cell(row=r, column=col["CNAE principal"]).value = rec.get("cnae_fiscal")
            ws.cell(row=r, column=col["CNAE descricao"]).value = rec.get("cnae_fiscal_descricao")
            ws.cell(row=r, column=col["Porte"]).value = rec.get("porte")
            ws.cell(row=r, column=col["Matriz/Filial"]).value = rec.get("descricao_identificador_matriz_filial")
            ws.cell(row=r, column=col["Situacao cadastral"]).value = rec.get("descricao_situacao_cadastral")
            ws.cell(row=r, column=col["Municipio"]).value = rec.get("municipio")
            ws.cell(row=r, column=col["UF"]).value = rec.get("uf")
        elif rec:
            ws.cell(row=r, column=col["CNAE descricao"]).value = rec.get("_erro")

        veredito, motivo = _avalia_loja_fisica(rec, filiais)
        ws.cell(row=r, column=col["Provavel loja fisica"]).value = veredito
        ws.cell(row=r, column=col["Motivo loja fisica"]).value = motivo
        if veredito == "provavel SIM":
            com_loja += 1

        arroba = ws.cell(row=r, column=col["Instagram (@)"]).value
        if tem_chave and arroba:
            ws.cell(row=r, column=col["Seguidores Instagram"]).value = _seguidores_instagram(str(arroba))
            time.sleep(0.6)

    CACHE_CNPJ.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    wb.save(XLSX_SAIDA)
    print(f"[ok] {XLSX_SAIDA.name} atualizado. {consultados} CNPJs consultados, "
          f"{com_loja} marcas com 'provavel SIM' pra loja fisica.")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "preparar":
        preparar()
    elif cmd == "enriquecer":
        enriquecer()
    else:
        sys.exit(__doc__)
