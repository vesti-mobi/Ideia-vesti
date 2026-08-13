"""
PainelElisa - TAMANHO DA MARCA (insumo pra oferta de upgrade de plano).

Le a planilha `Mensalidade até 400 .xlsx` (abas 'PRO' e 'Light  Starter '), busca
no Apify quantos seguidores a marca tem no Instagram e se ela aparece com loja
fisica no Google Maps, e grava:

  - aba 'Histórico' da propria planilha (1 linha por marca por coleta)
  - `tamanho_marca.json`, consumido por build_data.py -> aba "Tamanho da marca"

Por que isso importa: tem marca pagando plano antigo e barato que hoje e' grande
(muito seguidor, varias lojas). O painel cruza esse tamanho com a mensalidade
atual pra mostrar quem esta' mais defasado.

USO
    py tamanho_marca.py --limit 3      # piloto barato, so' 3 marcas
    py tamanho_marca.py                # roda tudo (pede confirmacao)
    py tamanho_marca.py --so-instagram # pula o Google Maps (bem mais barato)
    py tamanho_marca.py --so-json      # nao chama API: so' regera o JSON do historico

TOKEN
    Vem de APIFY_API_TOKEN ou do arquivo `.apify_token` (fora do git).
    NUNCA colocar o token no codigo: ele ja' vazou uma vez em gemini-code-*.txt.

CUSTO (plano FREE = US$ 5/mes de credito)
    Instagram e Maps sao chamados EM LOTE (uma run por bloco de marcas, nao uma
    run por marca) -- e' o que cabe no plano gratuito. O Maps e' o caro: por isso
    MAX_LOCAIS_POR_BUSCA limita quantos estabelecimentos ele coleta por marca.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT / "Mensalidade até 400 .xlsx"
OUT_JSON = ROOT / "tamanho_marca.json"
CACHE_CNPJ = ROOT / ".cache_cnpj.json"
# Cache do que ja' foi PAGO no Apify. Gravado a cada lote: se o script morrer no
# meio (aconteceu na 1a coleta cheia), nada precisa ser recomprado. Guarda os
# itens BRUTOS do Maps de proposito -- assim da' pra mudar o filtro
# _lugar_e_da_marca() e reprocessar de graca com --so-json.
CACHE_TAM = ROOT / ".cache_tamanho.json"
ABA_HISTORICO = "Histórico"

ATOR_INSTAGRAM = "apify~instagram-profile-scraper"
ATOR_MAPS = "compass~crawler-google-places"

LOTE_INSTAGRAM = 20          # usernames por run
LOTE_MAPS = 5                # buscas por run
MAX_LOCAIS_POR_BUSCA = 5     # teto de estabelecimentos por marca (custo do Maps)
PAUSA = 2                    # segundos entre runs

# CNAEs de varejo em loja -- reforcam "tem loja fisica" quando o Maps nao acha
CNAE_LOJA = ("4781", "4782", "4783", "4713", "4755")

# O Maps devolve QUALQUER estabelecimento que casa com o texto da busca. No piloto,
# "Gringa" trouxe Coxinha Da Gringa e Gringa Burger junto com a loja de roupa real.
# Sem estes dois filtros a coluna "tem loja fisica" vira lixo.
CATEGORIA_OK = ("roupa", "moda", "vestuario", "loja", "boutique", "confec", "calcado",
                "lingerie", "jeans", "atacado", "magazine", "acessorio", "bijuteria",
                "joalheria", "malha", "tecido", "fabricante", "textil", "shopping")
CATEGORIA_NAO = ("lanchonete", "restaurante", "sanduicheria", "pizzaria", "padaria",
                 "bar ", "hamburgueria", "hotel", "pousada", "salao", "barbearia",
                 "acougue", "mercado", "farmacia", "pet", "oficina", "auto", "borracharia",
                 "academia", "escola", "igreja", "clinica", "posto")
# palavras curtas/genericas nao servem pra casar nome de marca com nome de loja
STOPWORDS_MARCA = {"loja", "modas", "moda", "store", "brand", "confeccoes", "confeccao",
                   "ltda", "me", "eireli", "com", "de", "da", "do", "e", "&", "the",
                   "by", "oficial", "shop", "outlet", "jeans", "kids", "fashion"}


def _lugar_e_da_marca(marca: str, titulo: str, categoria: str) -> bool:
    """O estabelecimento do Maps e' mesmo desta marca?

    Exige as duas coisas: categoria de varejo/moda E o nome da marca aparecendo no
    nome do lugar. So' a categoria deixaria passar "MAFRA GRINGA"; so' o nome
    deixaria passar "Coxinha Da Gringa".
    """
    cat = _norm(categoria)
    if any(x in cat for x in CATEGORIA_NAO):
        return False
    if cat and not any(x in cat for x in CATEGORIA_OK):
        return False

    tit = _norm(titulo)
    termos = [t for t in re.split(r"[^a-z0-9]+", _norm(marca))
              if len(t) >= 3 and t not in STOPWORDS_MARCA]
    if not termos:                       # marca so' com palavra generica: cai no nome inteiro
        return _norm(marca) in tit
    return all(t in tit for t in termos) or sum(t in tit for t in termos) >= 2

INVALIDOS = {"", "desativado", "desativada", "nao tem", "não tem", "n/a", "na",
             "sem instagram", "-", "none", "null"}


# ---------------------------------------------------------------- utilidades
def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _valido(v: str) -> bool:
    return _txt(v).lower() not in INVALIDOS


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", _txt(s)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def _digitos(s) -> str:
    return re.sub(r"\D", "", _txt(s))


def _user_instagram(v: str) -> str:
    """'@marca', 'instagram.com/marca/' ou 'marca' -> 'marca'."""
    u = _txt(v)
    m = re.search(r"instagram\.com/([^/?#\s]+)", u, re.I)
    if m:
        u = m.group(1)
    return u.lstrip("@").strip().strip("/")


def _seguidores_para_int(v) -> int | None:
    """'133 mil' / '95,3 mil' / '1,2 mi' / '12345' -> int. None se nao der."""
    s = _norm(v).replace(".", "").replace("seguidores", "").strip()
    if not s:
        return None
    m = re.match(r"^([\d,]+)\s*(mil|mi|m|k)?$", s)
    if not m:
        return None
    try:
        num = float(m.group(1).replace(",", "."))
    except ValueError:
        return None
    mult = {"mil": 1_000, "k": 1_000, "mi": 1_000_000, "m": 1_000_000}.get(m.group(2) or "", 1)
    return int(num * mult)


def _token() -> str:
    tok = os.environ.get("APIFY_API_TOKEN", "").strip()
    if tok:
        return tok
    p = ROOT / ".apify_token"
    if p.exists():
        return p.read_text(encoding="utf-8").strip()
    sys.exit("[erro] sem token do Apify. Defina APIFY_API_TOKEN ou crie .apify_token")


def _get_json(url: str, timeout: int = 60):
    with urllib.request.urlopen(url, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _roda_ator(ator: str, entrada: dict, token: str, timeout: int = 300) -> list[dict]:
    """Roda o ator no modo sincrono. Bom para runs curtas (Instagram)."""
    url = (f"https://api.apify.com/v2/acts/{ator}/run-sync-get-dataset-items"
           f"?token={token}&timeout={timeout}")
    req = urllib.request.Request(
        url, data=json.dumps(entrada).encode("utf-8"),
        headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=timeout + 30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _roda_ator_async(ator: str, entrada: dict, token: str, espera_max: int = 900) -> list[dict]:
    """Dispara a run e fica perguntando ate' terminar.

    O Google Maps demora mais que o teto do endpoint `run-sync` (deu HTTP 408 no
    piloto), entao ele precisa deste caminho: POST /runs -> poll -> ler dataset.
    """
    req = urllib.request.Request(
        f"https://api.apify.com/v2/acts/{ator}/runs?token={token}",
        data=json.dumps(entrada).encode("utf-8"),
        headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=60) as resp:
        run = json.loads(resp.read().decode("utf-8"))["data"]

    run_id, dataset_id = run["id"], run["defaultDatasetId"]
    esperou = 0
    while esperou < espera_max:
        time.sleep(10)
        esperou += 10
        estado = _get_json(f"https://api.apify.com/v2/actor-runs/{run_id}?token={token}")["data"]
        status = estado.get("status")
        if status in ("SUCCEEDED", "FAILED", "TIMED-OUT", "ABORTED"):
            if status != "SUCCEEDED":
                raise RuntimeError(f"run {run_id} terminou como {status}")
            break
    else:
        raise TimeoutError(f"run {run_id} passou de {espera_max}s")

    return _get_json(
        f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={token}&clean=true",
        timeout=120)


# ---------------------------------------------------------------- planilha
ABAS = [
    ("PRO",             {"marca": "Marca",  "cnpj": "CNPJ", "insta": "@ Instagram",
                         "plano": "Plano",  "mensalidade": "Mensalidade", "cs": "CS"}),
    ("Light  Starter ", {"marca": "Marca",  "cnpj": "CNPJ", "insta": "@ Instagram",
                         "plano": "Plano",  "mensalidade": "Mensalidade", "canal": "Canal"}),
]


def _le_planilha() -> list[dict]:
    """Uma linha por marca das duas abas. Nomes de coluna com espaco sobressalente
    ('Marca ', '@ Instagram ') sao normalizados."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    marcas, vistos = [], set()
    for nome_aba, campos in ABAS:
        if nome_aba not in wb.sheetnames:
            print(f"[aviso] aba {nome_aba!r} nao existe na planilha", flush=True)
            continue
        ws = wb[nome_aba]
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            continue
        # mapeia nome-normalizado -> indice, ignorando espacos e caixa
        idx = {}
        for i, cab in enumerate(linhas[0]):
            chave = _norm(cab)
            if chave and chave not in idx:
                idx[chave] = i

        def val(linha, campo):
            i = idx.get(_norm(campos.get(campo, "")))
            return _txt(linha[i]) if i is not None and i < len(linha) else ""

        categoria = "PRO" if nome_aba == "PRO" else "Light Starter"
        for linha in linhas[1:]:
            if not any(x not in (None, "") for x in linha):
                continue
            marca = val(linha, "marca")
            if not _valido(marca):
                continue
            cnpj = _digitos(val(linha, "cnpj"))
            insta = _user_instagram(val(linha, "insta"))
            # regra da planilha: sem Instagram E sem CNPJ nao da' pra pesquisar nada
            if not _valido(insta) and len(cnpj) != 14:
                print(f"[pula] {marca}: sem Instagram e sem CNPJ validos", flush=True)
                continue
            chave = (cnpj or _norm(marca))
            if chave in vistos:
                continue
            vistos.add(chave)
            marcas.append({
                "categoria": categoria, "marca": marca, "cnpj": cnpj,
                "instagram": insta if _valido(insta) else "",
                "plano": val(linha, "plano"),
                "mensalidade": _seguidores_para_int(val(linha, "mensalidade")) or 0,
                "cs": val(linha, "cs"), "canal": val(linha, "canal"),
                "seguidoresPlanilha": _seguidores_para_int(val(linha, "seguidores")),
            })
    return marcas


def _cidade_do_cnpj(cnpj: str, cache: dict) -> str:
    """Municipio/UF do cache da BrasilAPI (gerado por upgrade_planilha.py).
    Melhora MUITO a busca no Maps -- 'Marca Sao Paulo SP' acha; so' 'Marca' nao."""
    rec = cache.get(cnpj) or {}
    if not isinstance(rec, dict) or rec.get("_erro"):
        return ""
    mun = _txt(rec.get("municipio"))
    uf = _txt(rec.get("uf"))
    return f"{mun} {uf}".strip()


# ---------------------------------------------------------------- cache
def _le_cache() -> dict:
    if CACHE_TAM.exists():
        try:
            c = json.loads(CACHE_TAM.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            c = {}
    else:
        c = {}
    c.setdefault("instagram", {})       # username -> dados do perfil
    c.setdefault("maps_brutos", {})     # termo de busca -> [itens crus do Maps]
    return c


def _grava_cache(cache: dict) -> None:
    CACHE_TAM.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")


def _termo_busca(m: dict, cache_cnpj: dict) -> str:
    cidade = _cidade_do_cnpj(m["cnpj"], cache_cnpj) if m["cnpj"] else ""
    return f"{m['marca']} {cidade}".strip() if cidade else m["marca"]


# ---------------------------------------------------------------- coletas
def coleta_instagram(marcas: list[dict], token: str, cache: dict) -> dict[str, dict]:
    """{username: {seguidores, ...}}. Em lote, pulando quem ja' esta' no cache."""
    faltam = [m["instagram"] for m in marcas
              if m["instagram"] and _norm(m["instagram"]) not in cache["instagram"]]
    if not faltam:
        print(f"[instagram] tudo em cache ({len(cache['instagram'])} perfis)", flush=True)
        return cache["instagram"]

    for i in range(0, len(faltam), LOTE_INSTAGRAM):
        bloco = faltam[i:i + LOTE_INSTAGRAM]
        print(f"[instagram] lote {i//LOTE_INSTAGRAM + 1}: {len(bloco)} perfis", flush=True)
        try:
            itens = _roda_ator(ATOR_INSTAGRAM, {"usernames": bloco}, token)
        except Exception as e:
            print(f"[instagram] lote falhou ({type(e).__name__}: {str(e)[:120]}) "
                  f"-- segue para o proximo", flush=True)
            continue
        for it in itens:
            u = _norm(it.get("username") or "")
            if u:
                cache["instagram"][u] = {
                    "seguidores": it.get("followersCount"),
                    "seguindo": it.get("followsCount"),
                    "publicacoes": it.get("postsCount"),
                    "verificado": bool(it.get("verified")),
                    "nomeCompleto": _txt(it.get("fullName")),
                }
        _grava_cache(cache)             # grava a CADA lote: interrupcao nao perde nada
        time.sleep(PAUSA)
    return cache["instagram"]


def coleta_maps(marcas: list[dict], token: str, cache_cnpj: dict, cache: dict) -> dict:
    """Guarda os itens BRUTOS por termo de busca. Pula termo ja' coletado."""
    termos = [_termo_busca(m, cache_cnpj) for m in marcas]
    faltam = [t for t in dict.fromkeys(termos) if t not in cache["maps_brutos"]]
    print(f"[maps] {len(termos)-len(faltam)} em cache, {len(faltam)} a coletar", flush=True)

    for i in range(0, len(faltam), LOTE_MAPS):
        bloco = faltam[i:i + LOTE_MAPS]
        print(f"[maps] lote {i//LOTE_MAPS + 1}/{-(-len(faltam)//LOTE_MAPS)}: "
              f"{len(bloco)} buscas", flush=True)
        entrada = {
            "searchStringsArray": bloco,
            "maxCrawledPlacesPerSearch": MAX_LOCAIS_POR_BUSCA,
            "language": "pt-BR",
            "countryCode": "br",
            "skipClosedPlaces": True,
        }
        try:
            itens = _roda_ator_async(ATOR_MAPS, entrada, token)
        except Exception as e:
            print(f"[maps] lote falhou ({type(e).__name__}: {str(e)[:120]}) "
                  f"-- segue para o proximo", flush=True)
            continue
        for t in bloco:                 # termo sem resultado vira lista vazia (nao recoletar)
            cache["maps_brutos"].setdefault(t, [])
        for it in itens:
            t = _txt(it.get("searchString"))
            if t:
                cache["maps_brutos"].setdefault(t, []).append({
                    "title": _txt(it.get("title")),
                    "categoryName": _txt(it.get("categoryName")),
                    "address": _txt(it.get("address") or it.get("street")),
                })
        _grava_cache(cache)
        time.sleep(PAUSA)
    return cache["maps_brutos"]


def aplica_filtro_maps(marcas: list[dict], brutos: dict, cache_cnpj: dict) -> dict:
    """Itens crus -> {termo: {qtd, enderecos}}. Puro: roda offline, de graca."""
    out, descartados = {}, 0
    for m in marcas:
        termo = _termo_busca(m, cache_cnpj)
        slot = out.setdefault(termo, {"qtd": 0, "enderecos": []})
        for it in brutos.get(termo, []):
            if not _lugar_e_da_marca(m["marca"], it.get("title"), it.get("categoryName")):
                descartados += 1
                continue
            slot["qtd"] += 1
            # guarda NOME + endereco: nome igual nao prova que e' a marca
            # ("MAFRA GRINGA" e "Gringa Beach" passam no filtro), entao quem liga
            # pra marca precisa ver o que foi encontrado antes de usar o numero.
            nome, end = it.get("title") or "", it.get("address") or ""
            if len(slot["enderecos"]) < 3:
                slot["enderecos"].append(f"{nome} — {end}" if end else nome)
    if descartados:
        print(f"[maps] {descartados} resultados descartados (categoria ou nome nao batem)",
              flush=True)
    out["_por_marca"] = {m["marca"]: _termo_busca(m, cache_cnpj) for m in marcas}
    return out


def dispara_maps(marcas: list[dict], token: str, cache_cnpj: dict, cache: dict,
                 por_run: int = 32) -> None:
    """Dispara as runs do Maps e SAI, sem esperar terminar.

    A run vive no Apify, nao neste processo: se o script morrer (ou o terminal
    fechar), ela termina do mesmo jeito e o resultado e' recolhido depois com
    --recuperar. E' o caminho certo pra coleta longa -- esperar em foreground
    ja' foi morto duas vezes por limite de tempo.
    """
    termos = [_termo_busca(m, cache_cnpj) for m in marcas]
    faltam = [t for t in dict.fromkeys(termos) if t not in cache["maps_brutos"]]
    if not faltam:
        print("[disparar] nada faltando", flush=True)
        return
    print(f"[disparar] {len(faltam)} buscas em {-(-len(faltam)//por_run)} run(s)", flush=True)
    for i in range(0, len(faltam), por_run):
        bloco = faltam[i:i + por_run]
        entrada = {
            "searchStringsArray": bloco,
            "maxCrawledPlacesPerSearch": MAX_LOCAIS_POR_BUSCA,
            "language": "pt-BR",
            "countryCode": "br",
            "skipClosedPlaces": True,
        }
        req = urllib.request.Request(
            f"https://api.apify.com/v2/acts/{ATOR_MAPS}/runs?token={token}",
            data=json.dumps(entrada).encode("utf-8"),
            headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=60) as resp:
            run = json.loads(resp.read().decode("utf-8"))["data"]
        print(f"[disparar] run {run['id']} com {len(bloco)} buscas", flush=True)
        time.sleep(PAUSA)
    print("[disparar] rode `py tamanho_marca.py --recuperar AAAA-MM-DD --so-json` "
          "quando as runs terminarem", flush=True)


def recupera_do_apify(token: str, cache: dict, desde: str) -> None:
    """Puxa runs JA' PAGAS do Apify pro cache. Ler dataset e' de graca -- serve pra
    aproveitar coleta interrompida em vez de comprar de novo."""
    for ator, rotulo in ((ATOR_INSTAGRAM, "instagram"), (ATOR_MAPS, "maps")):
        try:
            runs = _get_json(f"https://api.apify.com/v2/acts/{ator}/runs"
                             f"?token={token}&desc=true&limit=50")["data"]["items"]
        except Exception as e:
            print(f"[recuperar] {rotulo}: {type(e).__name__} {str(e)[:100]}", flush=True)
            continue
        novos = 0
        for r in runs:
            if r.get("status") != "SUCCEEDED" or _txt(r.get("startedAt")) < desde:
                continue
            try:
                itens = _get_json(f"https://api.apify.com/v2/datasets/{r['defaultDatasetId']}"
                                  f"/items?token={token}&clean=true", timeout=120)
            except Exception:
                continue
            for it in itens:
                if ator == ATOR_INSTAGRAM:
                    u = _norm(it.get("username") or "")
                    if u and u not in cache["instagram"]:
                        cache["instagram"][u] = {
                            "seguidores": it.get("followersCount"),
                            "seguindo": it.get("followsCount"),
                            "publicacoes": it.get("postsCount"),
                            "verificado": bool(it.get("verified")),
                            "nomeCompleto": _txt(it.get("fullName")),
                        }
                        novos += 1
                else:
                    t = _txt(it.get("searchString"))
                    if not t:
                        continue
                    lista = cache["maps_brutos"].setdefault(t, [])
                    reg = {"title": _txt(it.get("title")),
                           "categoryName": _txt(it.get("categoryName")),
                           "address": _txt(it.get("address") or it.get("street"))}
                    if reg not in lista:
                        lista.append(reg)
                        novos += 1
        print(f"[recuperar] {rotulo}: +{novos} registros de runs ja' pagas", flush=True)
    _grava_cache(cache)


# ---------------------------------------------------------------- consolidacao
def _porte(seguidores: int | None, unidades: int) -> str:
    """Faixa de tamanho da marca. Serve pra ordenar a fila de upgrade.

    Sem seguidores, loja fisica SOZINHA nao leva ao topo. Dois motivos:
    `qtdUnidades` satura em MAX_LOCAIS_POR_BUSCA (5 quase sempre significa "bateu
    o teto da busca", nao "tem exatamente 5 lojas") e o Maps ainda deixa passar
    homonimo. Na 1a coleta isso pos Valentin's e The Lion como "Muito grande" com
    seguidores desconhecidos -- errado, e logo no topo da fila de ligacao.
    """
    if seguidores is None:
        if unidades >= 3:
            return "Grande"
        return "Media" if unidades >= 1 else "Indefinido"
    if seguidores >= 500_000 or unidades >= 5:
        return "Muito grande"
    if seguidores >= 100_000 or unidades >= 3:
        return "Grande"
    if seguidores >= 20_000 or unidades >= 1:
        return "Media"
    return "Pequena" if seguidores > 0 else "Indefinido"


def consolida(marcas: list[dict], insta: dict, maps: dict, cache_cnpj: dict) -> list[dict]:
    por_marca = maps.get("_por_marca", {})
    hoje = date.today().isoformat()
    linhas = []
    for m in marcas:
        i = insta.get(_norm(m["instagram"])) if m["instagram"] else None
        seguidores = (i or {}).get("seguidores")
        if seguidores is None:
            seguidores = m.get("seguidoresPlanilha")     # fallback: o que ja' estava na planilha

        termo = por_marca.get(m["marca"], "")
        mp = maps.get(termo) or {}
        unidades = int(mp.get("qtd") or 0)
        enderecos = mp.get("enderecos") or []

        if unidades > 0:
            tem_loja, obs = "Sim", "; ".join(enderecos)
        else:
            # o Maps nao achou: o CNAE da Receita ainda pode indicar loja
            rec = cache_cnpj.get(m["cnpj"]) or {}
            cnae = _txt(rec.get("cnae_fiscal")) if isinstance(rec, dict) else ""
            if cnae[:4] in CNAE_LOJA:
                tem_loja, obs = "Provável", f"CNAE {cnae} é varejo em loja (Maps não achou)"
            else:
                tem_loja, obs = "Não", "Apenas E-commerce / Sem loja no Maps"

        linhas.append({
            "dataColeta": hoje,
            "categoria": m["categoria"],
            "marca": m["marca"],
            "cnpj": m["cnpj"],
            "instagram": m["instagram"],
            "seguidores": seguidores,
            "verificado": (i or {}).get("verificado", False),
            "temLojaFisica": tem_loja,
            "qtdUnidades": unidades,
            # bateu o teto da busca: o numero real de lojas pode ser maior
            "unidadesNoTeto": unidades >= MAX_LOCAIS_POR_BUSCA,
            "seguidoresDesconhecidos": seguidores is None,
            "enderecoObs": obs,
            "plano": m["plano"],
            "mensalidade": m["mensalidade"],
            "cs": m["cs"],
            "canal": m.get("canal", ""),
            "porte": _porte(seguidores, unidades),
        })
    return linhas


def grava_historico(linhas: list[dict]) -> None:
    """Acrescenta na aba 'Histórico' (append-only: cada coleta vira um bloco novo)."""
    wb = openpyxl.load_workbook(XLSX)
    if ABA_HISTORICO in wb.sheetnames:
        ws = wb[ABA_HISTORICO]
    else:
        ws = wb.create_sheet(ABA_HISTORICO)
        ws.append(["Data Coleta", "Categoria", "Marca", "CNPJ", "Instagram", "Seguidores",
                   "Tem loja física?", "Qtd Unidades Encontradas", "Endereço / Observação"])
    # primeira linha vazia (a aba ja' vem so' com cabecalho)
    destino = ws.max_row + 1
    while destino > 2 and all(c.value in (None, "") for c in ws[destino - 1]):
        destino -= 1
    for n, l in enumerate(linhas):
        ws.cell(destino + n, 1, l["dataColeta"])
        ws.cell(destino + n, 2, l["categoria"])
        ws.cell(destino + n, 3, l["marca"])
        ws.cell(destino + n, 4, l["cnpj"])
        ws.cell(destino + n, 5, l["instagram"])
        ws.cell(destino + n, 6, l["seguidores"])
        ws.cell(destino + n, 7, l["temLojaFisica"])
        ws.cell(destino + n, 8, l["qtdUnidades"])
        ws.cell(destino + n, 9, l["enderecoObs"])
    wb.save(XLSX)
    print(f"[write] aba {ABA_HISTORICO!r}: {len(linhas)} linhas a partir da {destino}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="processa so' N marcas (piloto)")
    ap.add_argument("--so-instagram", action="store_true", help="pula o Google Maps")
    ap.add_argument("--so-json", action="store_true", help="nao chama API; so' regera o JSON")
    ap.add_argument("--sim", action="store_true", help="nao pergunta antes de gastar credito")
    ap.add_argument("--historico", action="store_true",
                    help="grava a aba Histórico mesmo com --so-json")
    ap.add_argument("--disparar", action="store_true",
                    help="dispara as runs do Maps e sai (recolhe depois com --recuperar)")
    ap.add_argument("--recuperar", metavar="AAAA-MM-DD", default="",
                    help="puxa runs ja' pagas do Apify pro cache antes de coletar")
    args = ap.parse_args()

    if not XLSX.exists():
        sys.exit(f"[erro] planilha nao encontrada: {XLSX}")
    cache_cnpj = json.loads(CACHE_CNPJ.read_text(encoding="utf-8")) if CACHE_CNPJ.exists() else {}

    marcas = _le_planilha()
    if args.limit:
        marcas = marcas[:args.limit]
    com_insta = sum(1 for m in marcas if m["instagram"])
    print(f"[planilha] {len(marcas)} marcas ({com_insta} com Instagram)")

    cache = _le_cache()
    if args.recuperar:
        recupera_do_apify(_token(), cache, args.recuperar)
    if args.disparar:
        dispara_maps(marcas, _token(), cache_cnpj, cache)
        return

    if not args.so_json:
        falta_i = sum(1 for m in marcas
                      if m["instagram"] and _norm(m["instagram"]) not in cache["instagram"])
        falta_m = 0 if args.so_instagram else len(
            [t for t in dict.fromkeys(_termo_busca(m, cache_cnpj) for m in marcas)
             if t not in cache["maps_brutos"]])
        if not falta_i and not falta_m:
            print("[cache] nada a coletar -- tudo ja' esta' em cache")
        elif not args.sim:
            print(f"\nIsso vai gastar credito do Apify (plano FREE):"
                  f"\n  Instagram: {-(-falta_i // LOTE_INSTAGRAM)} run(s) para {falta_i} perfis"
                  f"\n  Maps:      {-(-falta_m // LOTE_MAPS)} run(s) para {falta_m} buscas "
                  f"(ate {MAX_LOCAIS_POR_BUSCA} locais cada)"
                  f"\n  (o que ja' esta' em cache nao e' recomprado)")
            if input("\nContinuar? [s/N] ").strip().lower() not in ("s", "sim", "y"):
                sys.exit("cancelado")
        if falta_i or falta_m:
            token = _token()
            coleta_instagram(marcas, token, cache)
            if not args.so_instagram:
                coleta_maps(marcas, token, cache_cnpj, cache)

    insta = cache["instagram"]
    maps = aplica_filtro_maps(marcas, cache["maps_brutos"], cache_cnpj)
    print(f"[cache] {len(insta)} perfis · {len(cache['maps_brutos'])} buscas de Maps")

    linhas = consolida(marcas, insta, maps, cache_cnpj)
    if not args.so_json or args.historico:
        grava_historico(linhas)

    OUT_JSON.write_text(json.dumps(linhas, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[write] {OUT_JSON.name} ({len(linhas)} marcas)")
    portes = {}
    for l in linhas:
        portes[l["porte"]] = portes.get(l["porte"], 0) + 1
    print("[porte]", portes)


if __name__ == "__main__":
    main()
