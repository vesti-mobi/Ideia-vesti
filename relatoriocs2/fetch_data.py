# -*- coding: utf-8 -*-
"""
Relatorio CS 2 - fetcher.

Gera dashboard_data.js (window.CS2_DATA) a partir de 3 fontes:
  Aba 1 (PRO)   -> Google Sheet "Implementacao > CS Passagem de bastao", aba PRO.
                   Le TODAS as colunas + a COR de cada linha (verde=ativa,
                   vermelho=cancelada, sem cor=sem reuniao / precisa de contato).
  Aba 2 (Tino)  -> API allblue-tinindo-tracking (/login_days).
  Aba 3 (Upsell)-> xlsx "ranking_crescimento_upsell", abas BUSTO / LUANA / THA.

Fontes das planilhas:
  - Em CI (workflow): se a env GOOGLE_SA_JSON estiver setada, baixa as planilhas
    frescas via Google Drive API (a conta de servico precisa ter acesso de leitura
    as duas planilhas). Caso contrario usa os arquivos locais _sheet1.xlsx / _sheet3.xlsx.
  - Tino: env TINO_USER / TINO_PASS (default admin / valor do secret).

Uso:  py fetch_data.py
"""
import os, io, json, base64, datetime, unicodedata, urllib.request, urllib.error

HERE = os.path.dirname(os.path.abspath(__file__))
SHEET1_ID = "1-_S4mSmez4_ugh0UMlMtLY7ATdI4llRfw2Uafgypdro"   # Passagem de bastao (nativo)
SHEET3_ID = "1V27Eo5uqeUR_MBI7wfk2DU3CqA7mtUbG"              # ranking_crescimento_upsell.xlsx
TINO_URL  = "https://allblue-tinindo-tracking-667335277398.us-central1.run.app/login_days"

GREEN_HEXES = {"FFB6D7A8", "B6D7A8"}
RED_HEXES   = {"FFE06666", "E06666", "FFEA9999", "EA9999", "FFF4CCCC", "FFCC0000"}

# ---------------------------------------------------------------- helpers
def today():
    return datetime.date.today()

def as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def iso(d):
    return d.isoformat() if d else None

def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    s = s.replace("R$", "").replace(" ", "").replace(" ", "")
    s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn")

# ---------------------------------------------------------------- download (CI)
def download_google_files():
    """Baixa as duas planilhas frescas via Google Drive API usando GOOGLE_SA_JSON."""
    sa_raw = os.environ.get("GOOGLE_SA_JSON")
    if not sa_raw:
        return False
    from google.oauth2 import service_account          # type: ignore
    from googleapiclient.discovery import build          # type: ignore
    from googleapiclient.http import MediaIoBaseDownload  # type: ignore
    info = json.loads(sa_raw)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive.readonly"])
    drive = build("drive", "v3", credentials=creds)
    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def save(req, path):
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        open(path, "wb").write(buf.getvalue())

    save(drive.files().export_media(fileId=SHEET1_ID, mimeType=XLSX),
         os.path.join(HERE, "_sheet1.xlsx"))
    save(drive.files().get_media(fileId=SHEET3_ID),
         os.path.join(HERE, "_sheet3.xlsx"))
    print("[google] planilhas baixadas via service account")
    return True

# ---------------------------------------------------------------- Aba 1
def classify_color(fill):
    if not fill or not getattr(fill, "patternType", None):
        return "sem_reuniao"
    rgb = getattr(fill.fgColor, "rgb", None)
    if not isinstance(rgb, str):
        return "sem_reuniao"
    rgb = rgb.upper()
    if rgb in GREEN_HEXES:
        return "ativa"
    if rgb in RED_HEXES:
        return "cancelada"
    # heuristica por canal (ARGB) para tons proximos
    if len(rgb) == 8:
        r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
        if g > r and g > b and g > 120:
            return "ativa"
        if r > g and r > b and r > 150:
            return "cancelada"
    return "sem_reuniao"

def build_aba1():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(HERE, "_sheet1.xlsx"))
    ws = wb["PRO"]
    t = today()
    rows = []
    for r in range(2, ws.max_row + 1):
        marca = ws.cell(row=r, column=3).value
        if not marca or not str(marca).strip():
            continue
        entrada = as_date(ws.cell(row=r, column=1).value)
        data25  = as_date(ws.cell(row=r, column=6).value)
        obs_g   = ws.cell(row=r, column=7).value
        obs_h   = ws.cell(row=r, column=8).value
        obs = " | ".join(str(x).strip() for x in (obs_g, obs_h) if x and str(x).strip())
        status = classify_color(ws.cell(row=r, column=3).fill)
        # texto "CANCELADO" reforca cancelada mesmo se cor faltar
        if status != "cancelada" and "CANCEL" in strip_accents(obs).upper():
            status = "cancelada"
        rows.append({
            "marca": str(marca).strip(),
            "entrada": iso(entrada),
            "dias": (t - entrada).days if entrada else None,
            "implementador": (ws.cell(row=r, column=4).value or ""),
            "cs": (ws.cell(row=r, column=5).value or ""),
            "data25": iso(data25),
            "obs_orig": obs,
            "status": status,
        })
    print(f"[aba1] {len(rows)} marcas")
    return rows

# ---------------------------------------------------------------- Aba 2 (Tino)
def build_aba2():
    user = os.environ.get("TINO_USER", "admin")
    pw   = os.environ.get("TINO_PASS", "hGh180#y")
    req = urllib.request.Request(TINO_URL, headers={
        "X-Dashboard-User": user, "X-Dashboard-Password": pw})
    with urllib.request.urlopen(req, timeout=45) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    data = payload.get("data", payload) if isinstance(payload, dict) else payload
    t = today()
    rows = []
    for r in data:
        ll = as_date(r.get("last_login"))
        dias_sem = (t - ll).days if ll else None
        rows.append({
            "company": r.get("company", ""),
            "nome": prettify(r.get("company", "")),
            "created_at": r.get("created_at"),
            "last_login": iso(ll),
            "dias_sem_acesso": dias_sem,      # None = nunca acessou
            "login_days": r.get("login_days"),
            "status": r.get("status"),
        })
    print(f"[aba2] {len(rows)} marcas com Tino")
    return rows

def prettify(slug):
    return " ".join(w.capitalize() for w in str(slug).replace("_", " ").split())

# ---------------------------------------------------------------- Aba 3
# mapa de colunas (1-indexed) por aba; None = coluna ausente
TAB_MAP = {
    "BUSTO":    dict(cs_label="Busto",    header_row=1, empresa=1, cs=2, plano=3,
                     gmv_ant=4, gmv_atual=5, cresc_rs=6, cresc_pct=7,
                     mensalidade=8, obs=9, tino=10, vestipago=None),
    "LUANA":    dict(cs_label="Luana",    header_row=1, empresa=2, cs=3, plano=4,
                     gmv_ant=5, gmv_atual=6, cresc_rs=7, cresc_pct=8,
                     mensalidade=None, obs=9, tino=None, vestipago=None),
    "THA":      dict(cs_label="Thamiris", header_row=1, empresa=2, cs=3, plano=4,
                     gmv_ant=5, gmv_atual=6, cresc_rs=7, cresc_pct=8,
                     mensalidade=None, obs=11, tino=9, vestipago=10),
}

def cell(ws, r, c):
    return ws.cell(row=r, column=c).value if c else None

def row_color(ws, r, col):
    """Cor de fundo (hex #rrggbb) da celula empresa; '' se sem cor / branco."""
    f = ws.cell(row=r, column=col).fill
    if not f or not getattr(f, "patternType", None):
        return ""
    rgb = getattr(f.fgColor, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) != 8:
        return ""
    hexc = "#" + rgb[2:].upper()
    if hexc in ("#FFFFFF", "#000000"):
        return ""
    return hexc

def build_aba3():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(HERE, "_sheet3.xlsx"), data_only=True)
    rows = []
    for tab, m in TAB_MAP.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(m["header_row"] + 1, ws.max_row + 1):
            empresa = cell(ws, r, m["empresa"])
            if not empresa or not str(empresa).strip() or str(empresa).strip().lower().startswith("empresa"):
                continue
            if not isinstance(empresa, str):
                continue
            gmv_ant = num(cell(ws, r, m["gmv_ant"]))
            gmv_at  = num(cell(ws, r, m["gmv_atual"]))
            cresc_rs = num(cell(ws, r, m["cresc_rs"]))
            if cresc_rs is None and gmv_at is not None and gmv_ant is not None:
                cresc_rs = gmv_at - gmv_ant
            pct = None
            if gmv_ant and gmv_ant > 0 and cresc_rs is not None:
                pct = cresc_rs / gmv_ant * 100.0
            def boolish(v):
                if v is None: return None
                if isinstance(v, bool): return v
                s = strip_accents(str(v)).strip().lower()
                if s in ("ok", "sim", "true", "1", "x", "ativo"): return True
                if s in ("nao", "false", "0", "-"): return False
                return str(v).strip()
            rows.append({
                "cs_tab": m["cs_label"],
                "empresa": str(empresa).strip(),
                "color": row_color(ws, r, m["empresa"]),   # cor original da planilha
                "cs": cell(ws, r, m["cs"]) or "",
                "plano": cell(ws, r, m["plano"]) or "",
                "gmv_ant": gmv_ant,
                "gmv_atual": gmv_at,
                "cresc_rs": cresc_rs,
                "cresc_pct": pct,
                "mensalidade": cell(ws, r, m["mensalidade"]) if m["mensalidade"] else None,
                "tino": boolish(cell(ws, r, m["tino"])) if m["tino"] else None,
                "vestipago": boolish(cell(ws, r, m["vestipago"])) if m["vestipago"] else None,
                "obs_orig": (str(cell(ws, r, m["obs"])).strip() if cell(ws, r, m["obs"]) else ""),
            })
    print(f"[aba3] {len(rows)} linhas (Busto/Luana/Thamiris)")
    return rows

# ---------------------------------------------------------------- main
def load_prev():
    """Le o dashboard_data.js atual (p/ preservar aba1/aba3 qdo nao ha acesso as planilhas)."""
    p = os.path.join(HERE, "dashboard_data.js")
    if not os.path.exists(p):
        return {}
    txt = open(p, encoding="utf-8").read().strip()
    txt = txt[txt.find("{"): txt.rfind("}") + 1]
    try:
        return json.loads(txt)
    except Exception:
        return {}

def main():
    has_sheets = False
    try:
        has_sheets = download_google_files()
    except Exception as e:
        print(f"[google] sem service account, tento arquivos locais ({e})")
    if not has_sheets:
        has_sheets = os.path.exists(os.path.join(HERE, "_sheet1.xlsx")) and \
                     os.path.exists(os.path.join(HERE, "_sheet3.xlsx"))

    prev = load_prev()
    if has_sheets:
        aba1, aba3 = build_aba1(), build_aba3()
    else:
        aba1 = prev.get("aba1", [])
        aba3 = prev.get("aba3", [])
        print(f"[aba1/aba3] planilhas indisponiveis — preservando snapshot "
              f"({len(aba1)}/{len(aba3)} linhas). Configure GOOGLE_SA_JSON p/ refresh automatico.")

    data = {
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "hoje": iso(today()),
        "aba1": aba1,
        "aba2": build_aba2(),   # Tino sempre atualiza (so precisa das credenciais da API)
        "aba3": aba3,
    }
    out = os.path.join(HERE, "dashboard_data.js")
    with open(out, "w", encoding="utf-8") as f:
        f.write("window.CS2_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=1)
        f.write(";\n")
    print(f"[ok] {out} gerado")

if __name__ == "__main__":
    main()
